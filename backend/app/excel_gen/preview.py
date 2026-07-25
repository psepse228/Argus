"""Converts a real generated Справка .xlsx into a plain-JSON grid the
frontend renders as an actual styled table -- a real "HUD" over the real
file (fonts, borders, merges, column widths, number formats), not a
separate hand-built numeric summary. Reads whatever the workbook actually
contains, so it can never drift from what "Скачать" would give you.
"""
import io
from datetime import datetime

import openpyxl
from openpyxl.utils import range_boundaries

# Real number formats used by writer.py/styles.py -- not a general Excel
# format parser, just enough to render these specific templates faithfully.
def _format_value(value, number_format: str) -> str:
    if value is None:
        return ""
    fmt = (number_format or "General").strip()
    try:
        if fmt == "0%":
            return f"{round(float(value) * 100)}%"
        if fmt in ("0.00",) or "#,##0.00" in fmt:
            return f"{float(value):,.2f}"
        if fmt == "mm-dd-yy":
            if isinstance(value, datetime):
                return value.strftime("%m-%d-%y")
            return str(value)
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    except (TypeError, ValueError):
        return str(value)


def _col_width_px(width: float | None) -> int:
    # openpyxl width is in "characters" of the default font -- ~7px/char plus
    # padding is the standard approximation used by most xlsx->HTML tools.
    if not width:
        return 64
    return round(width * 7 + 5)


def _border_css(border) -> dict:
    def side(s):
        if s is None or s.style is None:
            return "none"
        weight = "2px" if s.style == "medium" else "1px"
        return f"{weight} solid rgba(255,255,255,0.35)"
    return {
        "top": side(border.top), "bottom": side(border.bottom),
        "left": side(border.left), "right": side(border.right),
    }


def workbook_bytes_to_preview(data: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active

    merges = list(ws.merged_cells.ranges)
    # cells covered by a merge (other than its own top-left anchor) are
    # skipped entirely when rendering -- HTML expresses the merge via
    # colspan/rowspan on the anchor cell instead.
    covered: set[tuple[int, int]] = set()
    merge_at: dict[tuple[int, int], tuple[int, int]] = {}
    for rng in merges:
        min_col, min_row, max_col, max_row = range_boundaries(str(rng))
        merge_at[(min_row, min_col)] = (max_row - min_row + 1, max_col - min_col + 1)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if (r, c) != (min_row, min_col):
                    covered.add((r, c))

    max_row = ws.max_row
    max_col = ws.max_column

    col_widths = [_col_width_px(ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width) for c in range(1, max_col + 1)]
    row_heights = [round((ws.row_dimensions[r].height or 15) * 1.2) for r in range(1, max_row + 1)]

    rows = []
    for r in range(1, max_row + 1):
        cells = []
        for c in range(1, max_col + 1):
            if (r, c) in covered:
                continue
            cell = ws.cell(row=r, column=c)
            span = merge_at.get((r, c), (1, 1))
            font = cell.font
            cells.append({
                "value": _format_value(cell.value, cell.number_format),
                "rowSpan": span[0], "colSpan": span[1],
                "bold": bool(font.bold), "fontFamily": font.name or "inherit",
                "fontSize": round((font.size or 11) * 1.0),
                "align": (cell.alignment.horizontal or "left"),
                "wrap": bool(cell.alignment.wrap_text),
                "border": _border_css(cell.border),
            })
        rows.append(cells)

    return {"rows": rows, "colWidths": col_widths, "rowHeights": row_heights}
