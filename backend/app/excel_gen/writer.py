"""Writes a Справка .xlsx matching the real Italiano Vero template layout —
data (verified exact against real samples) and visual formatting (column
widths, row heights, borders, fonts, number formats — see styles.py for the
real fingerprint this was built from)."""
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Font

from .models import SpravkaInput
from .calc import SpravkaResult, compute
from . import styles as S

# Columns Q-W in the header row use the Arial header font in the real files;
# X (the last column) uses a distinct "last column" font. Everything else
# uses the Times New Roman header font. 1-indexed.
_HEADER_ARIAL_COLS = set(range(17, 24))  # Q..W
_HEADER_LAST_COL = 24  # X

# In the data row, columns B..J use the bold Arial data font, K..V use the
# bold Times New Roman data font, W/X use the "last" font (Calibri).
_DATA_ARIAL_COLS = set(range(2, 11))  # B..J
_DATA_LAST_COLS = {23, 24}  # W, X


def _pct_label(pct: float) -> int:
    """0.3 -> 30, matching how the real files bake the % into the header text."""
    return round(pct * 100)


def _setup_sheet(ws, col_widths: dict, row_heights: dict, zoom: int):
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width
    for row, height in row_heights.items():
        ws.row_dimensions[row].height = height
    ws.sheet_view.zoomScale = zoom


def _style_header_cell(ws, row: int, col: int, n_cols: int, value):
    c = ws.cell(row=row, column=col, value=value)
    if col in _HEADER_ARIAL_COLS:
        c.font = S.HEADER_FONT_ARIAL
    elif col == _HEADER_LAST_COL or col >= n_cols:
        c.font = S.HEADER_FONT_LAST
    else:
        c.font = S.HEADER_FONT_TIMES
    c.alignment = S.CENTER_WRAP
    if col == 1:
        c.border = S.BORDER_HEADER_FIRST
    elif col >= n_cols:
        c.border = S.BORDER_HEADER_LAST
    else:
        c.border = S.BORDER_HEADER_MID
    return c


def _style_data_cell(ws, row: int, col: int, n_cols: int, value, number_format: str = "General"):
    c = ws.cell(row=row, column=col, value=value)
    if col in _DATA_LAST_COLS:
        c.font = S.DATA_FONT_LAST
    elif col in _DATA_ARIAL_COLS:
        c.font = S.DATA_FONT_ARIAL
    else:
        c.font = S.DATA_FONT_TIMES
    c.alignment = S.CENTER_WRAP
    c.number_format = number_format
    if col == 1:
        c.border = S.BORDER_DATA_FIRST
    elif col >= n_cols:
        c.border = S.BORDER_DATA_LAST
    else:
        c.border = S.BORDER_DATA_MID
    return c


def build_spravka(inp: SpravkaInput, out_path: str) -> SpravkaResult:
    res = compute(inp)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"

    title = f"Справка по {inp.building} {inp.client_name}   {inp.client_phone}"

    if inp.is_full_payment:
        _write_full_payment(ws, inp, res, title)
    elif inp.balloon_months:
        _write_balloon(ws, inp, res, title)
    else:
        _write_installment(ws, inp, res, title)

    wb.save(out_path)
    return res


def _base_columns():
    return [
        "№", "этаж", "квадратура",
        "сумма за 1 кв.м.\n($) без акции ", "сумма за 1 кв.м.\n(сум) без акции ",
        "общая сумма сум без акции", "общая сумма ($)без акции",
        "общая сумма (Суммах ) акции со скидкой ", "общая сумма ($) акции со скидкой ",
        "сумма за 1 кв.м.\n($) ", "сумма за 1 кв.м. (суммах)",
        "общая сумма сум", "общая сумма ($)",
    ]


_ROW2_REGULAR_FONT = Font(name="Calibri", size=11, bold=False)


def _write_title_row(ws, title, n_cols):
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=2, column=col)
        c.font = _ROW2_REGULAR_FONT
        c.border = S.BORDER_TITLE if col == 1 else S.BORDER_TITLE_MID
    ws["A2"] = title
    ws["A2"].font = S.TITLE_FONT
    ws["A2"].alignment = S.CENTER_WRAP


def _style_footer_rows(ws, n_cols: int, row7_start: int = 1, row8_start: int = 1):
    """Rows 7 (ИТОГО) and 8 (курс/Менеджер) in the real template are fully
    bordered/fonted across their width even where a given cell is blank —
    replicating that box look rather than leaving a bare row after column A."""
    for row in (7, 8):
        for col in range(row7_start if row == 7 else row8_start, n_cols + 1):
            c = ws.cell(row=row, column=col)
            if c.value is None:
                c.value = None  # keep blank, just style it
            c.font = S.DATA_FONT_LAST if col in _DATA_LAST_COLS else S.DATA_FONT_TIMES
            c.alignment = S.CENTER_WRAP
            if col == 1:
                c.border = S.BORDER_DATA_FIRST
            elif col >= n_cols:
                c.border = S.BORDER_DATA_LAST
            else:
                c.border = S.BORDER_DATA_MID


def _write_full_payment(ws, inp, res, title):
    n_cols = 24  # A..X
    _setup_sheet(ws, S.COL_WIDTHS, S.ROW_HEIGHTS, S.ZOOM)
    _write_title_row(ws, title, n_cols)

    headers = _base_columns() + [None] * 11
    headers[13] = "Оплата"          # column N
    headers[23] = "Срок рассрочки"  # column X
    for col, label in enumerate(headers, start=1):
        if label is None:
            continue
        _style_header_cell(ws, 3, col, n_cols, label)
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)

    row = 5
    vals = [
        inp.unit_number, inp.floor, inp.area_m2,
        inp.price_per_m2_usd_no_promo, round(inp.price_per_m2_usd_no_promo * inp.exchange_rate, 2),
        res.total_no_promo_sum, res.total_no_promo_usd, res.total_promo_sum, res.total_promo_usd,
        res.effective_price_per_m2_usd, res.effective_price_per_m2_sum,
        res.effective_total_sum, res.effective_total_usd, 1,
    ] + [None] * 9 + [res.payment_label]
    fmts = ["General", "General", "General", "0.00", "General", "0.00", "0.00", "0.00", "0.00",
            "0.00", "0.00", "0.00", "0.00", "0%"] + ["General"] * 9 + ["General"]
    for col, (v, fmt) in enumerate(zip(vals, fmts), start=1):
        if v is None:
            continue
        _style_data_cell(ws, row, col, n_cols, v, fmt)

    _style_footer_rows(ws, n_cols)
    ws["A7"] = "ИТОГО"
    ws.merge_cells("A7:B7")
    ws["A8"] = "курс по таблицу"
    ws.merge_cells("A8:B8")
    ws["C8"] = inp.exchange_rate
    ws["C8"].number_format = "0.00"
    ws["W8"] = "Менеджер"
    ws["X8"] = inp.manager_name


def _write_installment(ws, inp, res, title):
    n_cols = 24  # A..X
    _setup_sheet(ws, S.COL_WIDTHS, S.ROW_HEIGHTS, S.ZOOM)
    _write_title_row(ws, title, n_cols)

    pct = _pct_label(inp.down_payment_pct)
    rem_pct = 100 - pct
    headers = _base_columns() + [
        "Оплата",
        f"Перво-начальный взнос (сум){pct}%",
        f"Перво- начальный взнос ($) {pct}%",
        f"Остаток суммы (сум) {rem_pct}% ",
        f"Остаток суммы ($) {rem_pct}%",
        f"Ежемесячная оплата в ( суммах) {inp.installment_months} месяцев будет платить",
        f"Ежемесячная оплата в ($) {inp.installment_months} месяцев будет  платить  ",
        f"оплата за {inp.installment_months} месяца итого (суммах )",
        f"оплата за {inp.installment_months} месяца итого ($ )",
        "Оплата ", "Срок рассрочки",
    ]
    for col, label in enumerate(headers, start=1):
        _style_header_cell(ws, 3, col, n_cols, label)
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)

    row = 5
    vals = [
        inp.unit_number, inp.floor, inp.area_m2,
        inp.price_per_m2_usd_no_promo, round(inp.price_per_m2_usd_no_promo * inp.exchange_rate, 2),
        res.total_no_promo_sum, res.total_no_promo_usd, res.total_promo_sum, res.total_promo_usd,
        res.effective_price_per_m2_usd, res.effective_price_per_m2_sum,
        res.effective_total_sum, res.effective_total_usd, res.payment_label,
        res.down_payment_sum, res.down_payment_usd, res.remaining_sum, res.remaining_usd,
        res.monthly_payment_sum, res.monthly_payment_usd,
        res.total_installment_paid_sum, res.total_installment_paid_usd,
    ]
    fmts = ["General", "General", "General", "0.00", "General", "0.00", "0.00", "0.00", "0.00",
            "0.00", "0.00", "0.00", "0.00", "0%", "0.00", "0.00", "0.00", "0.00", "0.00", "0.00", "0.00", "0.00"]
    for col, (v, fmt) in enumerate(zip(vals, fmts), start=1):
        _style_data_cell(ws, row, col, n_cols, v, fmt)
    start_dt = inp.payment_start_date or date.today()
    _style_data_cell(ws, row, 23, n_cols, datetime(start_dt.year, start_dt.month, start_dt.day), "mm-dd-yy")
    _style_data_cell(ws, row, 24, n_cols, inp.installment_months, "General")

    _style_footer_rows(ws, n_cols)
    ws["A7"] = "ИТОГО"
    ws.merge_cells("A7:B7")
    ws["A8"] = "курс по таблицу"
    ws.merge_cells("A8:B8")
    ws["C8"] = inp.exchange_rate
    ws["C8"].number_format = "0.00"
    ws["W8"] = "Менеджер"
    ws["X8"] = inp.manager_name


def _write_balloon(ws, inp, res, title):
    """Matches последний Task.xlsx's extra balloon-remainder columns. This
    real sample's own formatting differs slightly from the other two (all-
    thin borders, no medium outer box, taller header row) — replicated as
    observed, not assumed identical to the other two layouts."""
    n_cols = 25  # A..Y
    col_widths = {**S.COL_WIDTHS_BALLOON}
    row_heights = S.ROW_HEIGHTS_BALLOON
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width
    for r, h in row_heights.items():
        ws.row_dimensions[r].height = h
    ws.sheet_view.zoomScale = S.ZOOM_BALLOON

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws["A2"] = title
    ws["A2"].font = S.TITLE_FONT
    ws["A2"].alignment = S.CENTER_WRAP
    thin_box = S.BORDER_DATA_MID  # balloon layout uses thin borders throughout, even at the outer edge
    for col in range(1, n_cols + 1):
        ws.cell(row=2, column=col).border = thin_box

    pct = _pct_label(inp.down_payment_pct)
    rem_pct = 100 - pct
    headers = _base_columns() + [
        "Оплата",
        f"Перво-начальный взнос (сум){pct}%",
        "Перво- начальный взнос ($) %",
        f"Остаток суммы (сум) {rem_pct}% ",
        f"Остаток суммы ($) {rem_pct}%",
        f"Ежемесячная оплата в ( суммах) {inp.balloon_months} месяцев будет платить",
        f"Ежемесячная оплата в ($) {inp.balloon_months}месяцев будет  платить  ",
        f"оплата за {inp.balloon_months} месяцев итого (суммах )",
        f"оплата за {inp.balloon_months} месяца итого ($ )",
        f"Остаток полсе {inp.balloon_months} месяцев оплат ($)",
        f"Остаток после {inp.balloon_months} месяцев оплат (сум)",
        "Оплата ",
    ]
    for col, label in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=label)
        c.font = S.HEADER_FONT_ARIAL if col in _HEADER_ARIAL_COLS else S.HEADER_FONT_TIMES
        c.alignment = S.CENTER_WRAP
        c.border = thin_box
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)

    row = 5
    vals = [
        inp.unit_number, inp.floor, inp.area_m2,
        inp.price_per_m2_usd_no_promo, round(inp.price_per_m2_usd_no_promo * inp.exchange_rate, 2),
        res.total_no_promo_sum, res.total_no_promo_usd, res.total_promo_sum, res.total_promo_usd,
        res.effective_price_per_m2_usd, res.effective_price_per_m2_sum,
        res.effective_total_sum, res.effective_total_usd, res.payment_label,
        res.down_payment_sum, res.down_payment_usd, res.remaining_sum, res.remaining_usd,
        res.monthly_payment_sum, res.monthly_payment_usd,
        res.total_installment_paid_sum, res.total_installment_paid_usd,
        res.balloon_remaining_usd, res.balloon_remaining_sum,
    ]
    for col, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = S.DATA_FONT_ARIAL if col in _DATA_ARIAL_COLS else S.DATA_FONT_TIMES
        c.alignment = S.CENTER_WRAP
        c.number_format = "0.00" if col not in (1, 2, 14) else "General"
        c.border = thin_box

    ws["A7"] = "ИТОГО"
    ws.merge_cells("A7:B7")
    ws["AC7"] = "Менеджер"
    ws["AD7"] = f" {inp.manager_name}"
    ws["A8"] = "курс по таблицу"
    ws.merge_cells("A8:B8")
    ws["C8"] = inp.exchange_rate
    ws["C8"].number_format = "0.00"
    start_dt = inp.payment_start_date or date.today()
    ws.cell(row=4, column=28, value=datetime(start_dt.year, start_dt.month, start_dt.day))